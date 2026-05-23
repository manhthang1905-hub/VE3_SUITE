"""
repair_srt_text.py
==================
Backfill cot srt_text trong scenes va director_plan sheets
dua tren timestamp matching voi file SRT thuc te.

Cach dung:
    python repair_srt_text.py --excel <path_xlsx> --srt <path_srt>

Vi du:
    python repair_srt_text.py --excel "KA5-0072_prompts.xlsx" --srt "KA5-0073.srt"
"""
import sys
import re
import argparse
import shutil
from pathlib import Path
from datetime import timedelta

sys.path.insert(0, str(Path(__file__).parent))


# ─── Parse SRT file ──────────────────────────────────────────────────────────

def parse_time_str(t: str) -> float:
    """'00:01:23,456' hoặc '00:01:23.456' → seconds (float)"""
    t = t.strip().replace(',', '.')
    parts = t.split(':')
    h, m, s = int(parts[0]), int(parts[1]), float(parts[2])
    return h * 3600 + m * 60 + s


def parse_srt(path: Path):
    """Trả về list of (start_sec, end_sec, text)"""
    text = path.read_text(encoding='utf-8', errors='replace')
    pattern = re.compile(
        r'\d+\s*\n'
        r'(\d{2}:\d{2}:\d{2}[,\.]\d{3})\s*-->\s*(\d{2}:\d{2}:\d{2}[,\.]\d{3})\s*\n'
        r'([\s\S]*?)(?=\n\s*\n|\Z)',
        re.MULTILINE
    )
    entries = []
    for m in pattern.finditer(text):
        s = parse_time_str(m.group(1))
        e = parse_time_str(m.group(2))
        t = m.group(3).strip().replace('\n', ' ')
        if t:
            entries.append((s, e, t))
    return entries


# ─── Timestamp helpers ────────────────────────────────────────────────────────

def ts_to_sec(ts):
    """'00:01:23,456' → float seconds. Handles None gracefully."""
    if ts is None:
        return None
    ts = str(ts).strip()
    if not ts:
        return None
    try:
        return parse_time_str(ts)
    except Exception:
        return None


def find_srt_text(srt_entries, start_sec, end_sec, tolerance=0.5):
    """
    Tìm tất cả SRT entries có timestamp OVERLAP với [start_sec, end_sec].
    Returns: concatenated text string.
    """
    if start_sec is None or end_sec is None:
        return ""
    
    matched = []
    for (s, e, t) in srt_entries:
        # Overlap check: entry overlaps scene window (với tolerance)
        overlap_start = max(s, start_sec - tolerance)
        overlap_end   = min(e, end_sec + tolerance)
        if overlap_end > overlap_start:
            matched.append(t)
    
    return " ".join(matched)


# ─── Main ─────────────────────────────────────────────────────────────────────

def repair(excel_path: Path, srt_path: Path):
    from openpyxl import load_workbook

    print(f"Excel : {excel_path}")
    print(f"SRT   : {srt_path}")

    # Parse SRT
    print(f"\nParsing SRT...")
    srt_entries = parse_srt(srt_path)
    print(f"  Found {len(srt_entries)} SRT entries")
    if not srt_entries:
        print("[ERROR] No SRT entries found – check file format!")
        return

    # Backup
    bak = excel_path.with_suffix('.bak.xlsx')
    shutil.copy2(excel_path, bak)
    print(f"  Backup: {bak}")

    # Load workbook
    wb = load_workbook(excel_path)
    updated_scenes = 0
    updated_dp = 0

    # ── Fix 'scenes' sheet ────────────────────────────────────────────────────
    if 'scenes' in wb.sheetnames:
        ws = wb['scenes']
        
        # Tìm column index cho srt_text, srt_start, srt_end từ header
        headers = [c.value for c in ws[1]]
        try:
            col_srt_text  = headers.index('srt_text')  + 1   # 1-indexed
            col_srt_start = headers.index('srt_start') + 1
            col_srt_end   = headers.index('srt_end')   + 1
        except ValueError as e:
            print(f"  [WARN] scenes: column not found: {e}")
            col_srt_text  = 6   # fallback by position
            col_srt_start = 2
            col_srt_end   = 3

        print(f"\nRepairing 'scenes' sheet ({ws.max_row-1} rows)...")
        for row_idx in range(2, ws.max_row + 1):
            # Skip empty rows
            if ws.cell(row_idx, 1).value is None:
                continue
            
            current_srt = ws.cell(row_idx, col_srt_text).value
            if current_srt and str(current_srt).strip():
                continue  # Already has content, skip
            
            start_ts = ws.cell(row_idx, col_srt_start).value
            end_ts   = ws.cell(row_idx, col_srt_end).value
            
            start_sec = ts_to_sec(start_ts)
            end_sec   = ts_to_sec(end_ts)
            
            text = find_srt_text(srt_entries, start_sec, end_sec)
            if text:
                ws.cell(row_idx, col_srt_text).value = text[:1000]
                updated_scenes += 1
        
        print(f"  Updated {updated_scenes} scenes with srt_text")
    else:
        print("  [WARN] 'scenes' sheet not found")

    # ── Fix 'director_plan' sheet ─────────────────────────────────────────────
    if 'director_plan' in wb.sheetnames:
        ws_dp = wb['director_plan']
        
        headers_dp = [c.value for c in ws_dp[1]]
        try:
            col_dp_srt_text  = headers_dp.index('srt_text')  + 1
            col_dp_srt_start = headers_dp.index('srt_start') + 1
            col_dp_srt_end   = headers_dp.index('srt_end')   + 1
        except ValueError:
            col_dp_srt_text  = 5
            col_dp_srt_start = 2
            col_dp_srt_end   = 3

        print(f"\nRepairing 'director_plan' sheet ({ws_dp.max_row-1} rows)...")
        for row_idx in range(2, ws_dp.max_row + 1):
            if ws_dp.cell(row_idx, 1).value is None:
                continue
            
            current = ws_dp.cell(row_idx, col_dp_srt_text).value
            if current and str(current).strip():
                continue  # Already has content, skip
            
            start_ts = ws_dp.cell(row_idx, col_dp_srt_start).value
            end_ts   = ws_dp.cell(row_idx, col_dp_srt_end).value
            
            start_sec = ts_to_sec(start_ts)
            end_sec   = ts_to_sec(end_ts)
            
            text = find_srt_text(srt_entries, start_sec, end_sec)
            if text:
                ws_dp.cell(row_idx, col_dp_srt_text).value = text[:500]
                updated_dp += 1
        
        print(f"  Updated {updated_dp} director_plan rows with srt_text")

    # Save
    wb.save(excel_path)
    print(f"\n[SUCCESS] Saved to {excel_path}")
    print(f"  scenes updated : {updated_scenes}")
    print(f"  director_plan  : {updated_dp}")
    print(f"  Backup saved   : {bak}")


def auto_find_srt(excel_path: Path) -> Path:
    """Tự tìm SRT file trong cùng thư mục."""
    folder = excel_path.parent
    for f in folder.glob("*.srt"):
        return f
    return None


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Repair srt_text in scenes/director_plan sheets")
    parser.add_argument("--excel", required=True, help="Path to _prompts.xlsx")
    parser.add_argument("--srt",   default=None,  help="Path to .srt file (auto-detect if omitted)")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    if not excel_path.exists():
        print(f"[ERROR] Excel not found: {excel_path}")
        sys.exit(1)

    srt_path = Path(args.srt) if args.srt else auto_find_srt(excel_path)
    if not srt_path or not srt_path.exists():
        print(f"[ERROR] SRT file not found. Use --srt <path>")
        sys.exit(1)

    repair(excel_path, srt_path)
