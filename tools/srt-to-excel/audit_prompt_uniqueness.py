"""
Quick audit script to check prompt uniqueness in generated Excel files.
Usage: python audit_prompt_uniqueness.py PROJECTS/TL1-0065/TL1-0065_prompts.xlsx
"""

import sys
import re
from pathlib import Path
import openpyxl

# Fix Windows encoding
if sys.platform == "win32":
    if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        except:
            pass


def remove_style_boilerplate(prompt: str) -> str:
    """Remove common style boilerplate to focus on unique content."""
    text = str(prompt or "").lower()
    boilerplate_phrases = [
        "dark gray streetwear cool-flat psychology illustration",
        "matching the provided reference character",
        "provided reference character",
        "use the provided reference image",
        "preserve the reference character exactly",
        "reference character style",
        "channel art style",
        "psychology illustration",
        "clean illustration",
        "no real humans",
        "no photorealism",
        "no readable text",
        "no text",
        "no watermark",
        "flat design",
        "paper texture",
        "muted cool tones",
        "dark charcoal gray",
        "cool gray-lavender background",
        "soft oval shadow",
        "clean dark outline",
        "style",
        "palette",
        "character",
        "nv1",
    ]
    for phrase in boilerplate_phrases:
        text = text.replace(phrase, " ")
    # Remove image references
    text = re.sub(r"\([^)]*\.(?:png|jpg|jpeg|webp)\)", " ", text)
    # Remove visual concept labels
    text = re.sub(r"\b(?:visual focus|scene elements|body language|emotional tone|visual metaphor)\s*:", " ", text)
    # Keep only alphanumeric and Korean
    text = re.sub(r"[^a-z0-9가-힣]+", " ", text)
    # Remove stopwords
    stopwords = {
        "the", "and", "with", "for", "from", "into", "that", "this", "show", "shows", "showing",
        "through", "while", "clean", "soft", "simple", "image", "scene", "visual", "focus", "tone",
        "elements", "body", "language", "emotional", "metaphor", "prompt", "must", "exact", "identity",
    }
    return " ".join(token for token in text.split() if len(token) > 2 and token not in stopwords)


def calculate_similarity(prompt_a: str, prompt_b: str) -> float:
    """Calculate Jaccard similarity after removing boilerplate."""
    clean_a = remove_style_boilerplate(prompt_a)
    clean_b = remove_style_boilerplate(prompt_b)
    words_a = set(clean_a.split())
    words_b = set(clean_b.split())
    if not words_a or not words_b:
        return 0.0
    return len(words_a & words_b) / len(words_a | words_b)


def audit_excel(excel_path: str):
    """Audit prompt uniqueness in Excel file."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if "scenes" not in wb.sheetnames:
        print(f"ERROR: No 'scenes' sheet found in {excel_path}")
        return

    ws = wb["scenes"]
    headers = [cell.value for cell in ws[1]]

    # Find columns
    scene_id_col = headers.index("scene_id") + 1 if "scene_id" in headers else None
    img_prompt_col = headers.index("img_prompt") + 1 if "img_prompt" in headers else None
    srt_text_col = headers.index("srt_text") + 1 if "srt_text" in headers else None

    if not scene_id_col or not img_prompt_col:
        print("ERROR: Missing required columns (scene_id, img_prompt)")
        return

    # Read scenes
    scenes = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        scene_id = row[scene_id_col - 1].value
        img_prompt = row[img_prompt_col - 1].value
        srt_text = row[srt_text_col - 1].value if srt_text_col else ""
        if scene_id and img_prompt:
            scenes.append({
                "scene_id": scene_id,
                "img_prompt": str(img_prompt),
                "srt_text": str(srt_text or "")[:80],
            })

    print(f"\n{'='*80}")
    print(f"PROMPT UNIQUENESS AUDIT: {Path(excel_path).name}")
    print(f"{'='*80}")
    print(f"Total scenes: {len(scenes)}\n")

    # Check similarity between consecutive scenes
    high_similarity = []
    for i in range(len(scenes) - 1):
        scene_a = scenes[i]
        scene_b = scenes[i + 1]
        similarity = calculate_similarity(scene_a["img_prompt"], scene_b["img_prompt"])

        if similarity > 0.85:
            high_similarity.append({
                "scene_a": scene_a["scene_id"],
                "scene_b": scene_b["scene_id"],
                "similarity": similarity,
                "srt_a": scene_a["srt_text"],
                "srt_b": scene_b["srt_text"],
            })

    if high_similarity:
        print(f"WARNING: FOUND {len(high_similarity)} HIGH SIMILARITY PAIRS (>85%):\n")
        for issue in high_similarity[:10]:  # Show first 10
            print(f"  Scene {issue['scene_a']} <-> Scene {issue['scene_b']}: {issue['similarity']:.2%}")
            print(f"    SRT A: {issue['srt_a']}")
            print(f"    SRT B: {issue['srt_b']}")
            print()
    else:
        print("OK: ALL PROMPTS ARE SUFFICIENTLY UNIQUE (<85% similarity)\n")

    # Sample first 5 prompts to show structure
    print(f"\n{'='*80}")
    print("SAMPLE PROMPTS (first 3 scenes):")
    print(f"{'='*80}\n")
    for scene in scenes[:3]:
        print(f"Scene {scene['scene_id']}:")
        print(f"  SRT: {scene['srt_text']}")
        print(f"  Prompt: {scene['img_prompt'][:200]}...")
        print()

    # Check if visual concept fields exist in scene_planning
    if "scene_planning" in wb.sheetnames:
        ws_plan = wb["scene_planning"]
        plan_headers = [cell.value for cell in ws_plan[1]]
        visual_concept_fields = [
            "visual_focus", "visual_metaphor", "concrete_props",
            "body_language_key", "emotional_visual"
        ]
        found_fields = [f for f in visual_concept_fields if f in plan_headers]

        print(f"\n{'='*80}")
        print("VISUAL CONCEPT FIELDS IN scene_planning:")
        print(f"{'='*80}")
        if found_fields:
            print(f"OK: Found {len(found_fields)}/{len(visual_concept_fields)} fields: {', '.join(found_fields)}")
        else:
            print(f"WARNING: No visual concept fields found. Expected: {', '.join(visual_concept_fields)}")

    print(f"\n{'='*80}\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python audit_prompt_uniqueness.py <excel_file>")
        sys.exit(1)

    excel_path = sys.argv[1]
    if not Path(excel_path).exists():
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    audit_excel(excel_path)
