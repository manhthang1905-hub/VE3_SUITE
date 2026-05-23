from __future__ import annotations

import argparse
import contextlib
import io
import json
from collections import Counter
from pathlib import Path
from typing import Dict, Tuple

import openpyxl

from modules.prompt_quality import (
    check_psychology_prompt_quality,
    check_unsupported_prompt_details,
    normalize_psychology_style_profile,
    postprocess_img_prompt,
    postprocess_video_prompt,
    score_psychology_sample_style_video,
    configure_prompt_quality_ai,
)


SAFE_MOTION_TERMS = [
    "static frame",
    "very slow push-in",
    "very slow pull-back",
    "gentle pan left",
    "gentle pan right",
]


def resolve_channel_from_project(project_code: str) -> str:
    import re

    text = str(project_code or "").strip()
    match = re.match(r"^([A-Za-z]+\d+)-0*(\d+)$", text, flags=re.IGNORECASE)
    if match:
        return f"{match.group(1).upper()}-T{int(match.group(2))}"
    if re.match(r"^[A-Za-z]+\d+-T\d+$", text, flags=re.IGNORECASE):
        return text.upper()
    return text


def _ref_dir_for_topic(topic: str) -> str:
    return {"psychology": "psychology", "finance": "finance"}.get(topic, "psychology")


def load_psychology_style_profile(workbook_path: Path) -> Dict[str, str]:
    channel = resolve_channel_from_project(workbook_path.parent.name)
    topic = ""
    metadata_path = workbook_path.parent / ".nguon_runtime_metadata.yaml"
    if metadata_path.exists():
        import yaml

        metadata = yaml.safe_load(metadata_path.read_text(encoding="utf-8")) or {}
        channel = str(metadata.get("reference_channel") or channel).strip()
        topic = str(metadata.get("topic") or "").strip().lower()

    ref_dir = _ref_dir_for_topic(topic) if topic else "psychology"
    style_path = Path(__file__).resolve().parent / "reference_characters" / ref_dir / channel / "style.yaml"
    if not style_path.exists() and ref_dir == "psychology":
        style_path = Path(__file__).resolve().parent / "reference_characters" / "finance" / channel / "style.yaml"
    data = {}
    if style_path.exists():
        import yaml

        data = yaml.safe_load(style_path.read_text(encoding="utf-8")) or {}
    profile = normalize_psychology_style_profile(data)
    if data.get("deepseek_api_key"):
        configure_prompt_quality_ai(str(data.get("deepseek_api_key", "")), str(data.get("deepseek_model", "") or "deepseek-chat"))
    profile["resolved_channel"] = channel
    profile["style_source"] = str(style_path) if style_path.exists() else ""
    is_styled = bool(
        style_path.exists()
        or channel.upper().startswith("TL1-T") or channel.upper().startswith("TH1-T")
        or "psychology" in topic or "tâm lý" in topic or "tam ly" in topic
        or "finance" in topic or "tài chính" in topic or "tai chinh" in topic
    )
    profile["is_psychology_project"] = is_styled
    return profile


def audit_scene_metrics(headers: Dict[str, int], rows, style_profile: Dict[str, str] | None = None) -> Dict[str, object]:
    metrics = {
        "scene_count": 0,
        "img_unsupported_hits": 0,
        "psychology_quality_hits": 0,
        "video_inanimate_motion_hits": 0,
        "static_frame_count": 0,
        "motion_counter": Counter(),
        "top_unsupported_terms": Counter(),
        "top_psychology_quality_terms": Counter(),
        "scene_kind_counter": Counter(),
        "video_quality_missing": Counter(),
        "video_quality_critical": Counter(),
        "image_grounded_video_hits": 0,
        "strict_negative_video_hits": 0,
        "narration_beat_video_hits": 0,
    }
    for row in rows:
        metrics["scene_count"] += 1
        srt_text = str(row[headers["srt_text"]] or "")
        img_prompt = str(row[headers["img_prompt"]] or "")
        video_prompt = str(row[headers["video_prompt"]] or "")
        scene_kind = str(row[headers["scene_kind"]] or "").strip().lower()
        subject_mode = str(row[headers["subject_mode"]] or "").strip().lower()
        primary_subject = str(row[headers["primary_subject"]] or "")
        primary_action = str(row[headers["primary_action"]] or "")
        visual_anchor = str(row[headers["visual_anchor"]] or "")
        metrics["scene_kind_counter"][scene_kind] += 1

        unsupported = check_unsupported_prompt_details(
            srt_text=srt_text,
            img_prompt=img_prompt,
            primary_subject=primary_subject,
            primary_action=primary_action,
            visual_anchor=visual_anchor,
        )
        if unsupported:
            metrics["img_unsupported_hits"] += 1
            for term in unsupported:
                metrics["top_unsupported_terms"][term] += 1

        characters_cell = row[headers["characters_used"]] if "characters_used" in headers else ""
        char_ids = [
            cid.strip()
            for cid in str(characters_cell or "").split(",")
            if cid.strip() and cid.strip() != "[]"
        ]
        if style_profile and str(style_profile.get("audience_language", "") or "").strip():
            quality_issues = check_psychology_prompt_quality(
                img_prompt,
                srt_text=srt_text,
                char_ids=char_ids,
                style_profile=style_profile,
            )
            if quality_issues:
                metrics["psychology_quality_hits"] += 1
                for issue in quality_issues:
                    metrics["top_psychology_quality_terms"][issue] += 1

        low_video = video_prompt.lower()
        video_score = score_psychology_sample_style_video(video_prompt, srt_text, style_profile)
        for item in video_score.get("missing", []):
            metrics["video_quality_missing"][item] += 1
        for item in video_score.get("critical_failures", []):
            metrics["video_quality_critical"][item] += 1
        if "image-derived visual base" in low_video or "exact illustrated setup" in low_video:
            metrics["image_grounded_video_hits"] += 1
        strict_terms = ["no readable text", "no letters", "no numbers", "no captions", "no signs", "no logos", "no watermark"]
        if all(term in low_video for term in strict_terms):
            metrics["strict_negative_video_hits"] += 1
        if "faithfully visualize this narration beat" in low_video:
            metrics["narration_beat_video_hits"] += 1
        if scene_kind in {"object_detail", "environment_story"} and any(
            token in low_video for token in ["displays", "shows", "protrudes", "rests", "lies", "glows", "flickers", "remains open", "stays on"]
        ):
            metrics["video_inanimate_motion_hits"] += 1

        for motion in SAFE_MOTION_TERMS:
            if motion in low_video:
                metrics["motion_counter"][motion] += 1
                if motion == "static frame":
                    metrics["static_frame_count"] += 1
                break

    metrics["motion_counter"] = dict(metrics["motion_counter"])
    metrics["scene_kind_counter"] = dict(metrics["scene_kind_counter"])
    metrics["video_quality_missing"] = dict(metrics["video_quality_missing"].most_common(12))
    metrics["video_quality_critical"] = dict(metrics["video_quality_critical"].most_common(12))
    metrics["top_unsupported_terms"] = dict(metrics["top_unsupported_terms"].most_common(12))
    metrics["top_psychology_quality_terms"] = dict(metrics["top_psychology_quality_terms"].most_common(12))
    return metrics


def normalize_workbook(path: Path, apply_changes: bool) -> Tuple[Dict[str, object], Dict[str, object]]:
    style_profile = load_psychology_style_profile(path)
    wb = openpyxl.load_workbook(path)
    ws = wb["scenes"]
    rows = list(ws.iter_rows(values_only=True))
    headers = {str(cell).strip(): idx for idx, cell in enumerate(rows[0]) if cell is not None}
    required = [
        "scene_id", "srt_text", "img_prompt", "video_prompt", "scene_kind",
        "subject_mode", "primary_subject", "primary_action", "visual_anchor",
        "duration", "location_used", "characters_used",
    ]
    missing = [key for key in required if key not in headers]
    if missing:
        raise RuntimeError(f"Workbook {path.name} missing headers: {missing}")

    before = audit_scene_metrics(headers, rows[1:], style_profile=style_profile)
    changed_img = 0
    changed_video = 0

    is_psychology = bool(style_profile.get("is_psychology_project"))

    for excel_row_idx, row in enumerate(rows[1:], start=2):
        srt_text = str(row[headers["srt_text"]] or "")
        img_prompt = str(row[headers["img_prompt"]] or "")
        video_prompt = str(row[headers["video_prompt"]] or "")
        scene_kind = str(row[headers["scene_kind"]] or "")
        subject_mode = str(row[headers["subject_mode"]] or "")
        primary_subject = str(row[headers["primary_subject"]] or "")
        primary_action = str(row[headers["primary_action"]] or "")
        visual_anchor = str(row[headers["visual_anchor"]] or "")
        duration = float(row[headers["duration"]] or 6.0)
        loc_id = str(row[headers["location_used"]] or "").strip()
        char_ids = [
            cid.strip()
            for cid in str(row[headers["characters_used"]] or "").split(",")
            if cid.strip() and cid.strip() != "[]"
        ]
        char_lookup = {cid: f"{cid}.png" for cid in char_ids}
        loc_lookup = {loc_id: f"{loc_id}.png"} if loc_id else {}

        with contextlib.redirect_stderr(io.StringIO()):
            new_img = postprocess_img_prompt(
                img_prompt,
                char_ids=char_ids,
                loc_id=loc_id,
                char_image_lookup=char_lookup,
                loc_image_lookup=loc_lookup,
                context_lock="",
                srt_text=srt_text,
                primary_subject=primary_subject,
                primary_action=primary_action,
                visual_anchor=visual_anchor,
                topic="psychology" if is_psychology else "story",
                style_profile=style_profile if is_psychology else None,
            )
        if new_img and new_img != img_prompt:
            changed_img += 1
            if apply_changes:
                ws.cell(excel_row_idx, headers["img_prompt"] + 1, new_img)

        new_video = postprocess_video_prompt(
            video_prompt,
            duration=duration,
            visual_moment="",
            srt_text=srt_text,
            camera="",
            mood="",
            scene_kind=scene_kind,
            subject_mode=subject_mode,
            primary_subject=primary_subject,
            primary_action=primary_action,
            visual_anchor=visual_anchor,
            shot_function="",
            sequence_role="",
            topic="psychology" if is_psychology else "story",
            style_profile=style_profile if is_psychology else None,
            img_prompt=new_img or img_prompt,
        )
        if new_video and new_video != video_prompt:
            changed_video += 1
            if apply_changes:
                ws.cell(excel_row_idx, headers["video_prompt"] + 1, new_video)

    if apply_changes and (changed_img or changed_video):
        wb.save(path)

    if apply_changes:
        wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows_after = list(wb2["scenes"].iter_rows(values_only=True))
        after = audit_scene_metrics(headers, rows_after[1:], style_profile=style_profile)
    else:
        after = before
    after["changed_img_prompts"] = changed_img
    after["changed_video_prompts"] = changed_video
    after["resolved_channel"] = style_profile.get("resolved_channel", "")
    after["audience_language"] = style_profile.get("audience_language", "")
    return before, after


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit and normalize prompt workbooks.")
    parser.add_argument("--root", default=r"D:\VE3_SUITE\PROJECTS", help="Projects root")
    parser.add_argument("--apply", action="store_true", help="Write normalized prompts back to workbook")
    args = parser.parse_args()

    root = Path(args.root)
    report = {}
    for workbook in sorted(root.glob("*/*_prompts.xlsx")):
        try:
            before, after = normalize_workbook(workbook, apply_changes=args.apply)
            report[workbook.parent.name] = {
                "path": str(workbook),
                "before": before,
                "after": after,
            }
        except Exception as exc:
            report[workbook.parent.name] = {
                "path": str(workbook),
                "error": str(exc),
            }

    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
