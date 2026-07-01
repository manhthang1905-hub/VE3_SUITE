"""
I2V POOL — Image-to-Video hàng loạt cho project thật (vd TL1-0550), đa-account rotation.
Mỗi scene: upload ảnh (per account) -> mediaId -> generate batchAsyncGenerateVideoReferenceImages
(curl_cffi chrome + WARP) -> poll/download.
Dùng account Copy (1..10).
"""
import os, sys, time, threading, queue, re, glob
from concurrent.futures import ThreadPoolExecutor
import openpyxl

import warp
import flow_client as fc
from chrome_factory import ChromeAccount

THROTTLE_K = 4
PER_ACCOUNT_MAX = 15
V_ASPECT = "VIDEO_ASPECT_RATIO_LANDSCAPE"   # ảnh 1376x768 = 16:9
I_ASPECT = "IMAGE_ASPECT_RATIO_LANDSCAPE"


def read_scenes(xlsx, limit=None):
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb["scenes"]; hdr = [c.value for c in ws[1]]; ix = {h: i for i, h in enumerate(hdr)}
    out = []
    for r in range(2, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        sid = row[ix["scene_id"]]; vp = row[ix["video_prompt"]]; ip = row[ix["img_path"]]
        if sid and vp and ip and os.path.exists(ip):
            out.append((int(sid), str(vp), ip))
    out.sort()
    return out[:limit] if limit else out


def accounts_1_to_10(base=r"D:\VE3_SUITE"):
    accs = []
    for n in range(1, 11):
        p = os.path.join(base, f"GoogleChromePortable - Copy ({n})", "GoogleChromePortable.exe")
        if os.path.exists(p): accs.append(p)
    return accs


class I2VPool:
    def __init__(self, scenes, out_dir, accounts, concurrency=3, base_port=9320, poll_threads=5):
        self.work = queue.Queue()
        for s in scenes: self.work.put(s)
        self.total = len(scenes)
        self.out_dir = out_dir; os.makedirs(out_dir, exist_ok=True)
        self.accounts = queue.Queue()
        for a in accounts: self.accounts.put(a)
        self.concurrency = concurrency; self.base_port = base_port; self.poll_threads = poll_threads
        self.jobs = queue.Queue()
        self.done = []; self.failed = []
        self._lock = threading.Lock(); self._stop = False; self._submitted = 0

    def account_worker(self, port):
        while not self._stop:
            if self.work.empty() or self._submitted >= self.total:
                return
            try: launcher = self.accounts.get_nowait()
            except queue.Empty: return
            label = (re.search(r"Copy \((\d+)\)", launcher) or [None, "?"])[1]
            acc = ChromeAccount(launcher, port)
            try:
                if not acc.launch(through_warp=False) or not acc.wait_ready(40):
                    print(f"[acc{label}] launch fail"); continue
                bearer = acc.bearer(); cookie = fc.labs_cookie_header(acc.cookies())
                sid_ctx = f";{int(time.time()*1000)}"
                project = acc.ensure_project()
                email = acc.email()
                if not project:
                    print(f"[acc{label}] {email} NO PROJECT -> skip"); acc.close(); continue
                print(f"[acc{label}] {email} project={project} -> producing I2V")
            except Exception as e:
                print(f"[acc{label}] setup err {e}");
                try: acc.close()
                except: pass
                continue

            consec = 0; made = 0
            while not self._stop and made < PER_ACCOUNT_MAX and not self.work.empty():
                try: scene_id, prompt, img_path = self.work.get_nowait()
                except queue.Empty: break
                # 1) upload ảnh -> mediaId (retry nhẹ)
                img_mid = None
                for _ in range(3):
                    img_mid, err = fc.upload_image(bearer, project, sid_ctx, img_path, I_ASPECT)
                    if img_mid: break
                    if err and err.startswith("401"): bearer = acc.bearer(); continue
                    time.sleep(1)
                if not img_mid:
                    print(f"[acc{label}] scene{scene_id} UPLOAD fail ({err}) -> requeue"); self.work.put((scene_id, prompt, img_path)); consec += 1
                    if consec >= THROTTLE_K: break
                    continue
                # 2) generate I2V (retry token)
                ok = False
                for _ in range(8):
                    tok = acc.mint_token()
                    if not tok: time.sleep(0.4); continue
                    payload = fc.build_payload(prompt, project, tok, (int(time.time())*100+scene_id) % 2_000_000,
                                               aspect=V_ASPECT, reference_media_id=img_mid)
                    kind, data = fc.generate(bearer, payload, url=fc.GEN_I2V)
                    if kind == "ok":
                        mid = (fc.operation_names(data) or [None])[0]
                        if mid:
                            self.jobs.put((mid, scene_id, cookie))
                            with self._lock: self._submitted += 1
                            print(f"[acc{label}] gen OK scene{scene_id} vid={mid[:8]} ({self._submitted}/{self.total})")
                        consec = 0; made += 1; ok = True; break
                    elif kind == "auth": bearer = acc.bearer(); continue
                    else: consec += 1
                if not ok:
                    self.work.put((scene_id, prompt, img_path)); consec += 1
                if consec >= THROTTLE_K:
                    print(f"[acc{label}] throttled -> rotate"); break
            try: acc.close()
            except: pass

    def poll_worker(self):
        while not self._stop:
            if len(self.done) + len(self.failed) >= self.total: return
            try: mid, scene_id, cookie = self.jobs.get(timeout=5)
            except queue.Empty:
                if self._submitted >= self.total and self.jobs.empty(): return
                continue
            dst = os.path.join(self.out_dir, f"{scene_id}.mp4")
            ok = False
            for _ in range(90):
                try:
                    n, url = fc.media_url_and_download(mid, cookie, dst)
                    print(f"[dl] scene{scene_id} {n} bytes -> {scene_id}.mp4")
                    with self._lock: self.done.append(dst); ok = True; break
                except Exception:
                    time.sleep(6)
            if not ok:
                with self._lock: self.failed.append(scene_id)
                print(f"[dl] scene{scene_id} timeout")

    def run(self):
        print(f"[i2v] WARP={warp.ensure_proxy_mode()} accounts={self.accounts.qsize()} scenes={self.total}")
        for _ in range(self.poll_threads):
            threading.Thread(target=self.poll_worker, daemon=True).start()
        t0 = time.time()
        with ThreadPoolExecutor(max_workers=self.concurrency) as ex:
            for w in range(self.concurrency):
                ex.submit(self.account_worker, self.base_port + w)
        while len(self.done) + len(self.failed) < self.total and time.time() - t0 < 5400:
            if self.work.empty() and self.jobs.empty() and self._submitted <= len(self.done) + len(self.failed):
                break
            time.sleep(3)
        self._stop = True
        dt = time.time() - t0
        print(f"\n=== I2V DONE: {len(self.done)} videos, {len(self.failed)} failed, "
              f"{self.total-len(self.done)-len(self.failed)} left in {dt:.0f}s (~{round(len(self.done)/dt*3600) if dt else 0}/h) ===")
        return self.done


if __name__ == "__main__":
    proj = r"D:\VE3_SUITE\PROJECTS\TL1-0550"
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 8
    conc = int(sys.argv[2]) if len(sys.argv) > 2 else 3
    scenes = read_scenes(os.path.join(proj, "TL1-0550_prompts.xlsx"), limit=n)
    print(f"scenes to render: {len(scenes)} (first={scenes[0][0]}..{scenes[-1][0]})")
    pool = I2VPool(scenes, os.path.join(proj, "vid_engine"), accounts_1_to_10(), concurrency=conc)
    pool.run()
